# tsukasamiyashita/dxfflowchartmiya/DxfFlowchartMiya-1559b301efcf479ed8934d5e23d0d8559520cac5/app.py
import sys
import os
import json
import uuid
import math
import unicodedata
import datetime
import xml.etree.ElementTree as ET
from PyQt6.QtWidgets import (QApplication, QMainWindow, QToolBar, QFileDialog, QMessageBox, QGraphicsItemGroup,
                             QColorDialog, QLabel, QWidget, QDialog, QVBoxLayout, QHBoxLayout, QGroupBox, 
                             QRadioButton, QComboBox, QDoubleSpinBox, QPushButton, QGraphicsItem, QGraphicsView, 
                             QProxyStyle, QStyle, QDockWidget)
from PyQt6.QtCore import Qt, QRectF, QPointF, QMarginsF
from PyQt6.QtGui import (QPen, QBrush, QColor, QPainter, QImage, QAction, QActionGroup,
                         QPageSize, QPageLayout, QUndoStack, QCursor, QIcon)
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog, QPrintPreviewWidget, QPrinterInfo
from PyQt6.QtSvg import QSvgGenerator

import qtawesome as qta
import qdarktheme
import networkx as nx
import openpyxl
import darkdetect

from graphics import (GRID_SIZE, SceneStateCommand, FlowchartView, 
                      NodeItem, WaypointItem, EdgeItem, FlowchartScene, 
                      clip_line_to_node, DxfFrameItem)
from dxf_io import export_dxf_file, import_dxf_as_items

class ToolTipDelayStyle(QProxyStyle):
    def styleHint(self, hint, option=None, widget=None, returnData=None):
        if hint == QStyle.StyleHint.SH_ToolTip_WakeUpDelay: return 200
        return super().styleHint(hint, option, widget, returnData)

class CustomPrintPreviewDialog(QDialog):
    def __init__(self, main_window, has_selection=False):
        super().__init__(main_window); self.main_window = main_window
        self.setWindowTitle("印刷プレビューと設定"); self.resize(1100, 750)
        self.printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        
        settings_layout = QVBoxLayout()
        printer_group = QGroupBox("プリンタ"); pr_layout = QVBoxLayout()
        self.printer_combo = QComboBox()
        for p in QPrinterInfo.availablePrinters(): self.printer_combo.addItem(p.printerName(), p)
        idx = self.printer_combo.findText(QPrinterInfo.defaultPrinter().printerName())
        if idx >= 0: self.printer_combo.setCurrentIndex(idx)
        pr_layout.addWidget(self.printer_combo); printer_group.setLayout(pr_layout); settings_layout.addWidget(printer_group)
        
        paper_group = QGroupBox("用紙設定"); pp_layout = QVBoxLayout()
        h_size = QHBoxLayout(); h_size.addWidget(QLabel("サイズ:"))
        self.paper_size_combo = QComboBox()
        for name, sz in [("A4", QPageSize.PageSizeId.A4), ("A3", QPageSize.PageSizeId.A3), ("B5", QPageSize.PageSizeId.B5), ("B4", QPageSize.PageSizeId.B4), ("Letter", QPageSize.PageSizeId.Letter)]:
            self.paper_size_combo.addItem(name, sz)
        h_size.addWidget(self.paper_size_combo); pp_layout.addLayout(h_size)
        
        h_ori = QHBoxLayout(); h_ori.addWidget(QLabel("向き:"))
        self.ori_portrait = QRadioButton("縦"); self.ori_landscape = QRadioButton("横")
        self.ori_portrait.setChecked(True); h_ori.addWidget(self.ori_portrait); h_ori.addWidget(self.ori_landscape); pp_layout.addLayout(h_ori)
        
        h_margin = QHBoxLayout(); h_margin.addWidget(QLabel("余白(mm):"))
        self.margin_spin = QDoubleSpinBox(); self.margin_spin.setRange(0, 100); self.margin_spin.setValue(10.0)
        h_margin.addWidget(self.margin_spin); pp_layout.addLayout(h_margin); paper_group.setLayout(pp_layout); settings_layout.addWidget(paper_group)
        
        range_group = QGroupBox("印刷範囲"); range_layout = QVBoxLayout()
        self.radio_all = QRadioButton("図面全体"); self.radio_view = QRadioButton("現在の表示範囲"); self.radio_sel = QRadioButton("選択したアイテム")
        self.radio_all.setChecked(True); self.radio_sel.setEnabled(has_selection)
        for r in [self.radio_all, self.radio_view, self.radio_sel]: range_layout.addWidget(r)
        range_group.setLayout(range_layout); settings_layout.addWidget(range_group)
        
        scale_group = QGroupBox("スケール設定"); sc_layout = QVBoxLayout()
        self.radio_auto = QRadioButton("自動調整"); self.radio_custom = QRadioButton("倍率指定(%)"); self.radio_auto.setChecked(True)
        h_scale = QHBoxLayout(); self.spin_scale = QDoubleSpinBox(); self.spin_scale.setRange(10, 1000); self.spin_scale.setValue(100); self.spin_scale.setEnabled(False)
        h_scale.addWidget(self.radio_custom); h_scale.addWidget(self.spin_scale); sc_layout.addWidget(self.radio_auto); sc_layout.addLayout(h_scale); scale_group.setLayout(sc_layout); settings_layout.addWidget(scale_group)
        
        btn_layout = QVBoxLayout(); self.btn_print = QPushButton("🖨️ 印刷を実行"); self.btn_print.setStyleSheet("font-weight: bold; padding: 10px;")
        self.btn_cancel = QPushButton("キャンセル"); btn_layout.addSpacing(20); btn_layout.addWidget(self.btn_print); btn_layout.addWidget(self.btn_cancel); settings_layout.addLayout(btn_layout)
        settings_layout.addStretch()
        
        self.preview_widget = QPrintPreviewWidget(self.printer); self.preview_widget.paintRequested.connect(self.handle_paint_request)
        main_layout = QHBoxLayout(self); left_panel = QWidget(); left_panel.setLayout(settings_layout); left_panel.setFixedWidth(280)
        main_layout.addWidget(left_panel); main_layout.addWidget(self.preview_widget, stretch=1)
        
        self.radio_custom.toggled.connect(self.spin_scale.setEnabled)
        for w in [self.printer_combo, self.paper_size_combo, self.ori_portrait, self.ori_landscape, self.margin_spin, self.radio_all, self.radio_view, self.radio_sel, self.radio_auto, self.spin_scale]:
            if isinstance(w, QComboBox): w.currentIndexChanged.connect(self.update_preview)
            elif isinstance(w, QRadioButton): w.toggled.connect(self.update_preview)
            elif isinstance(w, QDoubleSpinBox): w.valueChanged.connect(self.update_preview)
        self.btn_print.clicked.connect(self.do_print); self.btn_cancel.clicked.connect(self.reject)
        self.update_preview()

    def update_printer_settings(self):
        if self.printer_combo.currentData(): self.printer.setPrinterName(self.printer_combo.currentData().printerName())
        self.printer.setPageSize(QPageSize(self.paper_size_combo.currentData()))
        self.printer.setPageOrientation(QPageLayout.Orientation.Portrait if self.ori_portrait.isChecked() else QPageLayout.Orientation.Landscape)
        m = self.margin_spin.value(); self.printer.setPageMargins(QMarginsF(m, m, m, m), QPageLayout.Unit.Millimeter)

    def update_preview(self): self.update_printer_settings(); self.preview_widget.updatePreview()

    def do_print(self):
        self.update_printer_settings(); dialog = QPrintDialog(self.printer, self)
        if dialog.exec() == QPrintDialog.DialogCode.Accepted: self.handle_paint_request(self.printer); self.accept()

    def handle_paint_request(self, printer):
        if self.radio_all.isChecked(): print_rect = self.main_window.scene.itemsBoundingRect(); sel_only = False
        elif self.radio_view.isChecked(): print_rect = self.main_window.view.mapToScene(self.main_window.view.viewport().rect()).boundingRect(); sel_only = False
        else:
            rect = QRectF()
            for si in self.main_window.scene.selectedItems(): rect = rect.united(si.sceneBoundingRect())
            print_rect = rect; sel_only = True

        if print_rect.isEmpty(): return
        sel_items = self.main_window.scene.selectedItems(); self.main_window.scene.clearSelection()
        rect = QRectF(print_rect).adjusted(-5, -5, 5, 5); painter = QPainter(printer); painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        page_rect = printer.pageRect(QPrinter.Unit.DevicePixel); hidden_items = []
        
        def should_keep(it):
            if it in sel_items: return True
            for c in it.childItems():
                if should_keep(c): return True
            return False

        for item in self.main_window.scene.items():
            if not item.isVisible(): continue
            is_preview = False
            try:
                if item == getattr(self.main_window.scene, 'preview_node', None) or item in getattr(self.main_window.scene, 'preview_items', []):
                    is_preview = True
            except RuntimeError: pass

            if isinstance(item, WaypointItem) or is_preview or (sel_only and item.parentItem() is None and not should_keep(item)):
                item.hide(); hidden_items.append(item)

        old_grid = self.main_window.scene.draw_grid
        self.main_window.scene.draw_grid = False
        if self.radio_auto.isChecked(): self.main_window.scene.render(painter, QRectF(page_rect), rect, Qt.AspectRatioMode.KeepAspectRatio)
        else:
            sc = (self.spin_scale.value() / 100.0) * (printer.resolution() / self.main_window.logicalDpiX())
            sw, sh = rect.width() * sc, rect.height() * sc
            tx, ty = page_rect.left() + max(0, (page_rect.width()-sw)/2.0), page_rect.top() + max(0, (page_rect.height()-sh)/2.0)
            self.main_window.scene.render(painter, QRectF(tx, ty, sw, sh), rect, Qt.AspectRatioMode.KeepAspectRatio)
        
        self.main_window.scene.draw_grid = old_grid
        for item in hidden_items: item.show()
        painter.end()
        for item in sel_items: item.setSelected(True)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_filepath = None
        self.current_tool = "select"
        self.clipboard_data = None
        self.clipboard_base_pos = None
        self.undo_stack = QUndoStack(self)
        self.dxf_template_path = None
        self.dxf_template_scale = 1.0
        self.frame_item = None
        self.last_state = {"nodes": [], "edges": [], "groups": [], "dxf_template_path": None, "dxf_template_pos": None, "dxf_template_scale": 1.0}
        self.is_light_theme = not darkdetect.isDark()

        self.scene = FlowchartScene(self)
        self.scene.setSceneRect(-2000, -2000, 4000, 4000)

        self.view = FlowchartView(self.scene)
        self.setCentralWidget(self.view)
        
        icon_path = os.path.join(os.path.dirname(__file__), "icon.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
            
        self.init_legend()

        self.icon_actions = [] 
        self.init_menu()
        self.init_toolbars()
        self.load_settings()
        self.apply_theme() 
        
        self.scene.selectionChanged.connect(self.on_selection_changed)
        self.on_selection_changed() 
        
        self.update_window_title()
        self.statusBar().showMessage("準備完了: 範囲選択や複数選択（Ctrlキー+クリック）が可能です")

    def create_icon_action(self, icon_name, text, slot=None, shortcut=None, checkable=False):
        act = QAction(text, self)
        if shortcut: act.setShortcut(shortcut)
        if checkable: act.setCheckable(True)
        if slot: act.triggered.connect(slot)
        self.icon_actions.append((act, icon_name))
        return act

    def get_config_path(self):
        return os.path.join(os.path.expanduser("~"), "DxfFlowchartMiya", "settings.json")

    def save_settings(self):
        reply = QMessageBox.question(self, "設定保存の確認", "現在の書式設定（フォントや線のスタイルなど）を、次回起動時の既定値として保存しますか？",
                                   QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.No: return
        
        config_path = self.get_config_path()
        os.makedirs(os.path.dirname(config_path), exist_ok=True)
        
        settings = {
            "dxf_ver": self.cb_dxf_ver.currentIndex(),
            "font": self.cb_font.currentText(),
            "line_width": self.cb_width.currentText(),
            "line_style": self.cb_style.currentText(),
            "routing": self.cb_routing.currentIndex(),
            "arrow": self.cb_arrow.currentIndex(),
            "node_w": self.sb_node_w.value(),
            "node_h": self.sb_node_h.value(),
            "draw_grid": self.scene.draw_grid
        }
        
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=4, ensure_ascii=False)
            QMessageBox.information(self, "完了", f"設定を保存しました:\n{config_path}")
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"設定の保存に失敗しました:\n{e}")

    def load_settings(self):
        config_path = self.get_config_path()
        if not os.path.exists(config_path): return
        try:
            with open(config_path, 'r', encoding='utf-8') as f: settings = json.load(f)
            if "dxf_ver" in settings: self.cb_dxf_ver.setCurrentIndex(settings["dxf_ver"])
            if "font" in settings: self.cb_font.setCurrentText(settings["font"])
            if "line_width" in settings: self.cb_width.setCurrentText(settings["line_width"])
            if "line_style" in settings: self.cb_style.setCurrentText(settings["line_style"])
            if "routing" in settings: self.cb_routing.setCurrentIndex(settings["routing"])
            if "arrow" in settings: self.cb_arrow.setCurrentIndex(settings["arrow"])
            if "node_w" in settings: self.sb_node_w.setValue(settings["node_w"])
            if "node_h" in settings: self.sb_node_h.setValue(settings["node_h"])
            if "draw_grid" in settings: 
                self.scene.draw_grid = settings["draw_grid"]
                self.act_grid.setChecked(settings["draw_grid"])
        except Exception as e: print(f"Failed to load settings: {e}")

    def reset_settings(self):
        reply = QMessageBox.question(self, "設定リセットの確認", "保存されている既定の設定を削除し、アプリの初期状態に戻しますか？",
                                   QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.No: return
        config_path = self.get_config_path()
        if os.path.exists(config_path):
            try: os.remove(config_path)
            except Exception as e: print(f"Failed to delete settings file: {e}")
        
        self.cb_dxf_ver.setCurrentIndex(1)
        self.cb_font.setCurrentText("ＭＳ ゴシック")
        self.cb_width.setCurrentText("2")
        self.cb_style.setCurrentIndex(0) 
        self.cb_routing.setCurrentIndex(0) 
        self.cb_arrow.setCurrentIndex(1) 
        self.sb_node_w.setValue(100)
        self.sb_node_h.setValue(50)
        self.scene.draw_grid = True
        self.act_grid.setChecked(True)
        self.scene.update()
        QMessageBox.information(self, "リセット", "設定をデフォルトに戻しました。")

    def apply_theme(self):
        theme = "light" if self.is_light_theme else "dark"
        app = QApplication.instance()
        if app:
            try:
                if hasattr(qdarktheme, 'setup_theme'): qdarktheme.setup_theme(theme)
                else: app.setStyleSheet(qdarktheme.load_stylesheet(theme))
                app.setStyle(ToolTipDelayStyle(app.style()))
            except Exception as e: print(f"Theme setup failed: {e}")
        
        text_color = "#212529" if self.is_light_theme else "#F8F9FA"
        bg_hover = "#E2E6EA" if self.is_light_theme else "#495057"
        
        self.setStyleSheet(f"""
            QToolBar {{ spacing: 6px; padding: 4px; border: none; }} 
            QToolButton {{ font-size: 13px; padding: 6px 10px; border-radius: 4px; color: {text_color}; }} 
            QToolButton:hover {{ background: {bg_hover}; }}
            QToolButton:checked {{ background: #3B82F6; color: white; font-weight: bold; }} 
        """)

        icon_color = '#212529' if self.is_light_theme else '#F8F9FA'
        for act, icon_name in self.icon_actions:
            if icon_name: act.setIcon(qta.icon(icon_name, color=icon_color))

        if self.is_light_theme: self.scene.setBackgroundBrush(QBrush(QColor(255, 255, 255)))
        else: self.scene.setBackgroundBrush(QBrush(QColor(30, 30, 30)))
        
        for item in self.scene.items():
            if hasattr(item, 'update_pen'): item.update_pen()
            if hasattr(item, 'update_style'): item.update_style()
        self.scene.update()

    def toggle_grid(self):
        self.scene.draw_grid = not self.scene.draw_grid
        self.scene.update()

    def get_scene_json(self, selected_only=False):
        frame_pos = None
        if self.frame_item:
            frame_pos = {"x": self.frame_item.scenePos().x(), "y": self.frame_item.scenePos().y()}
        
        data = {"nodes": [], "edges": [], "groups": [], "dxf_template_path": self.dxf_template_path, "dxf_template_pos": frame_pos, "dxf_template_scale": self.dxf_template_scale}
        items_raw = self.scene.selectedItems() if selected_only else self.scene.items()
        
        items = set()
        for it in items_raw:
            items.add(it)
            if type(it) == QGraphicsItemGroup:
                for child in it.childItems(): items.add(child)
        items = list(items)
        valid_node_ids = set()
        
        for item in items:
            if isinstance(item, NodeItem) and getattr(self.scene, 'preview_node', None) != item and item not in getattr(self.scene, 'preview_items', []):
                data["nodes"].append({"id": item.node_id, "type": item.node_type, "x": item.scenePos().x(), "y": item.scenePos().y(), "w": item.w, "h": item.h, "text": item.text_item.toPlainText(), "bg_color": item.bg_color.name(), "text_color": item.text_color.name(), "line_color": item.line_color.name() if item.line_color else None, "font": item.font_family})
                valid_node_ids.add(item.node_id)
                
        for item in items:
            if isinstance(item, EdgeItem) and item not in getattr(self.scene, 'preview_items', []):
                if selected_only and (item.source_node.node_id not in valid_node_ids or item.target_node.node_id not in valid_node_ids): continue
                offset = {"x": item.text_item.manual_offset.x(), "y": item.text_item.manual_offset.y()} if item.text_item.manual_offset else None
                data["edges"].append({"source": item.source_node.node_id, "target": item.target_node.node_id, "label": item.raw_text, "width": item.line_width, "style": item.line_style, "routing": item.routing, "arrow": item.arrow, "line_color": item.line_color.name() if item.line_color else None, "font": item.font_family, "waypoints": [{"x": wp.scenePos().x(), "y": wp.scenePos().y()} for wp in item.waypoints], "text_offset": offset})
                
        for item in items:
            if type(item) == QGraphicsItemGroup:
                c_ids = [c.node_id for c in item.childItems() if hasattr(c, 'node_id')]
                if c_ids: data["groups"].append(c_ids)
                
        return data

    def load_scene_json(self, data, offset_x=0, offset_y=0, clear_scene=True, generate_new_ids=False, is_undo_redo=False):
        if clear_scene: 
            self.scene.clear(); self.scene.items_ref.clear()
            self.scene.preview_node = None; self.scene.preview_items = []; self.scene.source_node = None
            
        id_map = {}
        
        self.dxf_template_path = data.get("dxf_template_path")
        self.update_dxf_template_label()
        
        t_pos = data.get("dxf_template_pos")
        pos = QPointF(t_pos["x"], t_pos["y"]) if t_pos else QPointF(0, 0)
        self.dxf_template_scale = data.get("dxf_template_scale", 1.0)
        self.sb_frame_scale.blockSignals(True)
        self.sb_frame_scale.setValue(self.dxf_template_scale)
        self.sb_frame_scale.blockSignals(False)
        
        if self.dxf_template_path:
            self.load_dxf_frame(self.dxf_template_path, pos, self.dxf_template_scale)
        elif self.frame_item:
            self.scene.removeItem(self.frame_item)
            self.frame_item = None
            self.act_lock_frame.setEnabled(False)

        for n in data.get("nodes", []):
            new_id = str(uuid.uuid4()) if generate_new_ids else n.get("id")
            node = NodeItem(n["x"]+offset_x, n["y"]+offset_y, n["text"], n["type"], new_id, n.get("bg_color", "#E1F5FE"), n.get("text_color", "#000000"), w=n.get("w", 100), h=n.get("h", 50), line_color=n.get("line_color"))
            if n.get("font"): node.set_font_family(n.get("font"))
            self.scene.items_ref.append(node); self.scene.addItem(node); id_map[n.get("id")] = node
            if not clear_scene: node.setSelected(True)
            
        for e in data.get("edges", []):
            src, tgt = id_map.get(e["source"]), id_map.get(e["target"])
            if src and tgt:
                edge = EdgeItem(src, tgt, e.get("label", ""), e.get("width", 2), e.get("style", "solid"), e.get("routing", "straight"), e.get("arrow", "end"), line_color=e.get("line_color"))
                if e.get("font"): edge.set_font_family(e.get("font"))
                if e.get("text_offset"): edge.text_item.manual_offset = QPointF(e.get("text_offset")["x"], e.get("text_offset")["y"])
                for w in e.get("waypoints", []):
                    wp = WaypointItem(w["x"]+offset_x, w["y"]+offset_y, edge); edge.waypoints.append(wp); self.scene.items_ref.append(wp); self.scene.addItem(wp)
                src.add_edge(edge); tgt.add_edge(edge); self.scene.items_ref.append(edge); self.scene.addItem(edge); edge.update_position()
                if not clear_scene: edge.setSelected(True)
                
        for g_cids in data.get("groups", []):
            g_items = [id_map[cid] for cid in g_cids if cid in id_map]
            if g_items:
                group = self.scene.createItemGroup(g_items)
                group.setFlags(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable | QGraphicsItem.GraphicsItemFlag.ItemIsMovable)
                self.scene.items_ref.append(group); getattr(group, 'setSelected', lambda x: None)(True if not clear_scene else False)
                
        if not is_undo_redo: self.last_state = self.get_scene_json()

    def push_undo_state(self, description):
        new_state = self.get_scene_json()
        if self.last_state != new_state:
            self.undo_stack.push(SceneStateCommand(self, self.last_state, new_state, description))
            self.last_state = new_state

    def update_window_title(self):
        base = "DxfFlowchartMiya v1.0.0"
        self.setWindowTitle(f"{os.path.basename(self.current_filepath)} - {base}" if self.current_filepath else base)

    def init_menu(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu("ファイル(&F)")
        for text, func, sc in [("上書き保存(&S)", self.save_file, "Ctrl+S"), ("名前を付けて保存(&A)...", self.save_file_as, "Ctrl+Shift+S"), ("読込(&O)", self.load_json, "Ctrl+O")]:
            act = QAction(text, self); act.setShortcut(sc); act.triggered.connect(func); file_menu.addAction(act)
        file_menu.addSeparator()
        file_menu.addAction("DXF出力の図枠リンクを設定...", self.link_dxf_template)
        file_menu.addAction("DXF図枠リンクを解除", self.clear_dxf_template)
        file_menu.addSeparator()
        
        act_excel = self.create_icon_action('fa5s.file-excel', "仕様書(Excel)生成...", self.generate_excel)
        file_menu.addAction(act_excel)
        file_menu.addSeparator()
        file_menu.addAction("Draw.ioインポート(.xml)...", self.import_drawio)
        file_menu.addAction("Draw.ioエクスポート(.xml)...", self.export_drawio)
        file_menu.addSeparator()
        file_menu.addSeparator()
        file_menu.addAction("DXFエクスポート...", lambda: self.export_file("DXF Files (*.dxf)"))
        file_menu.addAction("PDFエクスポート...", lambda: self.export_file("PDF Files (*.pdf)"))
        file_menu.addAction("JPEGエクスポート...", lambda: self.export_file("JPEG Files (*.jpeg *.jpg)"))
        file_menu.addAction("その他の形式でエクスポート...", lambda: self.export_file())
        file_menu.addAction("Jw_cadへコピー(&C)", self.copy_to_jwcad)
        file_menu.addSeparator(); file_menu.addAction("終了(&X)", self.close)
        
        edit_menu = menubar.addMenu("編集(&E)")
        self.act_undo = self.undo_stack.createUndoAction(self, "元に戻す(&U)")
        self.act_undo.setShortcut("Ctrl+Z")
        self.icon_actions.append((self.act_undo, 'fa5s.undo'))
        edit_menu.addAction(self.act_undo)

        self.act_redo = self.undo_stack.createRedoAction(self, "やり直し(&R)")
        self.act_redo.setShortcut("Ctrl+Y")
        self.icon_actions.append((self.act_redo, 'fa5s.redo'))
        edit_menu.addAction(self.act_redo)
        edit_menu.addSeparator()
        
        act_copy = self.create_icon_action('fa5s.copy', "コピー(&C)", self.copy_items, shortcut="Ctrl+C")
        act_paste = self.create_icon_action('fa5s.paste', "貼り付け(&V)", self.paste_items, shortcut="Ctrl+V")
        act_del = self.create_icon_action('fa5s.trash-alt', "削除(&D)", self.delete_selected_items, shortcut="Del")
        edit_menu.addAction(act_copy); edit_menu.addAction(act_paste); edit_menu.addAction(act_del)
        edit_menu.addSeparator()
        
        act_grp = QAction("グループ化(&G)", self); act_grp.setShortcut("Ctrl+G"); act_grp.triggered.connect(self.group_selected); edit_menu.addAction(act_grp)
        act_ungrp = QAction("グループ解除(&U)", self); act_ungrp.setShortcut("Ctrl+Shift+G"); act_ungrp.triggered.connect(self.ungroup_selected); edit_menu.addAction(act_ungrp)
        
        arr_menu = menubar.addMenu("配置(&A)")
        act_layout = self.create_icon_action('fa5s.sitemap', "★自動階層レイアウト", self.auto_layout_networkx)
        arr_menu.addAction(act_layout)
        arr_menu.addSeparator()
        for txt, mode in [("左揃え", "left"), ("左右中央揃え", "center_x"), ("右揃え", "right"), ("上揃え", "top"), ("上下中央揃え", "center_y"), ("下揃え", "bottom"), ("水平等間隔", "dist_h"), ("垂直等間隔", "dist_v")]:
            act = QAction(txt, self); act.triggered.connect(lambda chk, m=mode: self.align_items(m)); arr_menu.addAction(act)

        view_menu = menubar.addMenu("表示(&V)")
        self.act_grid = self.create_icon_action('fa5s.th', "グリッド表示/非表示", self.toggle_grid, checkable=True)
        self.act_grid.setChecked(True); view_menu.addAction(self.act_grid)
        
        help_menu = menubar.addMenu("ヘルプ(&H)")
        help_menu.addAction("使い方(&U)", self.show_usage); help_menu.addAction("バージョン情報(&A)", self.show_about)

    def show_usage(self):
        msg = ("【操作説明】\n"
               "・図形の追従プレビュー: 追加モード時やコピペ時、カーソルにゴーストが追従します。\n"
               "・Undo/Redo: Ctrl+Z / Ctrl+Y\n"
               "・コピー＆ペースト: Ctrl+Cでコピーし、キャンバスをクリックして配置\n"
               "・グループ化: Ctrl+G / 解除: Ctrl+Shift+G\n"
               "・整列 / 自動レイアウト: 複数選択して上部メニューの「配置」から実行\n"
               "・線のスタイル: エッジを選択して「書式ツールバー」で太さや直角配線、矢印の有無を変更\n\n"
               "・DXF図枠読込 / Jw_cad連携 / 仕様書生成 / Draw.io互換 は「ファイル」メニューから実行可能です。")
        QMessageBox.information(self, "使い方", msg)

    def show_about(self): 
        QMessageBox.about(self, "情報", "DxfFlowchartMiya v1.0.0\nPython & PyQt6 製フローチャート作成ツール")

    def init_legend(self):
        dock = QDockWidget("ノード解説", self)
        dock.setFeatures(QDockWidget.DockWidgetFeature.DockWidgetMovable | QDockWidget.DockWidgetFeature.DockWidgetFloatable)
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        items = [
            ('fa5s.square', "処理 (Process)", "一般的な処理や工程、計算に使います。長方形で表します。"),
            ('fa5s.code-branch', "分岐 (Decision)", "条件によってYes/Noなどが分かれる判断に使います。菱形で表します。"),
            ('fa5s.layer-group', "データ (Data)", "データの入出力や情報の流れを示します。平行四辺形で表します。"),
            ('fa5s.capsules', "端子 (Terminal)", "フローの開始（スタート）や終了（エンド）を示します。角丸またはカプセル型で表します。")
        ]
        
        ic_color = 'gray'
        for icon, title, desc in items:
            row = QHBoxLayout()
            label_icon = QLabel()
            label_icon.setPixmap(qta.icon(icon, color=ic_color).pixmap(24, 24))
            row.addWidget(label_icon, 0, Qt.AlignmentFlag.AlignTop)
            
            txt_layout = QVBoxLayout()
            label_title = QLabel(f"<b>{title}</b>")
            label_desc = QLabel(desc)
            label_desc.setWordWrap(True)
            label_desc.setStyleSheet("font-size: 11px; color: gray;")
            txt_layout.addWidget(label_title)
            txt_layout.addWidget(label_desc)
            row.addLayout(txt_layout)
            layout.addLayout(row)
            layout.addSpacing(15)
        
        layout.addStretch()
        dock.setWidget(widget)
        self.addDockWidget(Qt.DockWidgetArea.RightDockWidgetArea, dock)

    def init_toolbars(self):
        tb_main = QToolBar("メインツール"); self.addToolBar(tb_main)
        self.action_group = QActionGroup(self)
        self.btn_select = self.create_icon_action('fa5s.mouse-pointer', "選択", lambda: self.set_tool("select"), checkable=True)
        self.btn_select.setChecked(True)
        tb_main.addAction(self.btn_select)
        self.action_group.addAction(self.btn_select)
        tb_main.addSeparator()
        
        for icon_name, text, key in [("fa5s.square", "処理", "process"), ("fa5s.code-branch", "分岐", "decision"), ("fa5s.layer-group", "データ", "data"), ("fa5s.capsules", "端子", "terminal")]:
            act = self.create_icon_action(icon_name, text, lambda chk, k=key: self.set_tool(k), checkable=True)
            tb_main.addAction(act); self.action_group.addAction(act)
        
        tb_main.addSeparator()
        act_conn = self.create_icon_action('fa5s.link', "接続", lambda: self.set_tool("connect"), checkable=True)
        tb_main.addAction(act_conn); self.action_group.addAction(act_conn)
        tb_main.addSeparator()
        tb_main.addAction(self.act_grid)

        tb_edit = QToolBar("編集"); self.addToolBar(tb_edit)
        tb_edit.addAction(self.act_undo)
        tb_edit.addAction(self.act_redo)
        tb_edit.addSeparator()
        for act, _ in self.icon_actions:
            if act.text() in ["コピー(&C)", "貼り付け(&V)", "削除(&D)"]:
                tb_edit.addAction(act)
        
        tb_cad = QToolBar("CAD設定"); self.addToolBar(tb_cad)
        self.lbl_dxf_template = QLabel(" 図枠リンク: [未設定] "); tb_cad.addWidget(self.lbl_dxf_template)
        tb_cad.addSeparator()
        tb_cad.addWidget(QLabel("DXF版:")); self.cb_dxf_ver = QComboBox()
        self.cb_dxf_ver.addItems(["2007", "2010", "2013", "2018"])
        self.cb_dxf_ver.setItemData(0, "R2007"); self.cb_dxf_ver.setItemData(1, "R2010")
        self.cb_dxf_ver.setItemData(2, "R2013"); self.cb_dxf_ver.setItemData(3, "R2018")
        self.cb_dxf_ver.setCurrentIndex(1); tb_cad.addWidget(self.cb_dxf_ver)
        
        tb_cad.addSeparator()
        self.act_lock_frame = self.create_icon_action('fa5s.lock', "図面枠をロック中 (クリックで解除)", self.toggle_frame_lock, checkable=True)
        self.act_lock_frame.setChecked(True)
        self.act_lock_frame.setEnabled(False)
        tb_cad.addAction(self.act_lock_frame)
        
        tb_cad.addSeparator()
        tb_cad.addWidget(QLabel(" 図枠倍率: "))
        self.sb_frame_scale = QDoubleSpinBox()
        self.sb_frame_scale.setRange(0.01, 100.0)
        self.sb_frame_scale.setSingleStep(0.1)
        self.sb_frame_scale.setValue(1.0)
        self.sb_frame_scale.setEnabled(False)
        self.sb_frame_scale.valueChanged.connect(self.change_frame_scale)
        tb_cad.addWidget(self.sb_frame_scale)
        
        self.addToolBarBreak()

        tb_style = QToolBar("書式"); self.addToolBar(tb_style)
        act_bg = self.create_icon_action('fa5s.fill-drip', "背景", self.change_bg_color)
        act_fg = self.create_icon_action('fa5s.font', "文字色", self.change_text_color)
        act_ln = self.create_icon_action('fa5s.pen', "線の色", self.change_line_color)
        tb_style.addActions([act_bg, act_fg, act_ln])
        tb_style.addSeparator()
        tb_style.addWidget(QLabel("フォント:")); self.cb_font = QComboBox(); self.cb_font.addItems(["ＭＳ ゴシック", "標準SHXフォント", "unicode"]); self.cb_font.currentTextChanged.connect(self.change_font_family); tb_style.addWidget(self.cb_font)
        tb_style.addSeparator()
        tb_style.addWidget(QLabel("線幅:")); self.cb_width = QComboBox(); self.cb_width.addItems(["1", "2", "3", "4", "5"]); self.cb_width.setCurrentText("2"); self.cb_width.currentTextChanged.connect(self.change_edge_style); tb_style.addWidget(self.cb_width)
        tb_style.addWidget(QLabel("線種:")); self.cb_style = QComboBox(); self.cb_style.addItems(["実線(solid)", "破線(dash)", "点線(dot)"]); self.cb_style.currentTextChanged.connect(self.change_edge_style); tb_style.addWidget(self.cb_style)
        tb_style.addWidget(QLabel("ルート:")); self.cb_routing = QComboBox(); self.cb_routing.addItems(["直線", "直角(Orthogonal)"]); self.cb_routing.setItemData(0, "straight"); self.cb_routing.setItemData(1, "orthogonal"); self.cb_routing.currentIndexChanged.connect(self.change_edge_style); tb_style.addWidget(self.cb_routing)
        tb_style.addWidget(QLabel("終端:")); self.cb_arrow = QComboBox(); self.cb_arrow.addItems(["なし", "矢印(終端)", "矢印(始端)", "両矢印"]); self.cb_arrow.setItemData(0, "none"); self.cb_arrow.setItemData(1, "end"); self.cb_arrow.setItemData(2, "start"); self.cb_arrow.setItemData(3, "both"); self.cb_arrow.setCurrentIndex(1); self.cb_arrow.currentIndexChanged.connect(self.change_edge_style); tb_style.addWidget(self.cb_arrow)
        tb_style.addSeparator()
        tb_style.addWidget(QLabel("幅:")); self.sb_node_w = QDoubleSpinBox(); self.sb_node_w.setRange(20, 1000); self.sb_node_w.setSingleStep(20); self.sb_node_w.setValue(100); self.sb_node_w.valueChanged.connect(self.on_node_size_ui_changed); tb_style.addWidget(self.sb_node_w)
        tb_style.addWidget(QLabel("高さ:")); self.sb_node_h = QDoubleSpinBox(); self.sb_node_h.setRange(20, 1000); self.sb_node_h.setSingleStep(20); self.sb_node_h.setValue(50); self.sb_node_h.valueChanged.connect(self.on_node_size_ui_changed); tb_style.addWidget(self.sb_node_h)
        tb_style.addSeparator()
        act_save_cfg = self.create_icon_action('fa5s.save', "規定値として設定保存", self.save_settings)
        tb_style.addAction(act_save_cfg)
        act_reset_cfg = self.create_icon_action('fa5s.undo-alt', "設定をデフォルトに戻す", self.reset_settings)
        tb_style.addAction(act_reset_cfg)

    def set_tool(self, tool_name):
        self.current_tool = tool_name
        if self.scene.source_node: self.scene.source_node.set_highlight(False); self.scene.source_node = None
        self.scene.clearSelection()
        
        if tool_name == "select": 
            self.view.setDragMode(QGraphicsView.DragMode.RubberBandDrag); self.view.setCursor(Qt.CursorShape.ArrowCursor); self.statusBar().showMessage("準備完了")
            self.scene.update_preview_node(None, tool_name)
        elif tool_name == "connect": 
            self.view.setDragMode(QGraphicsView.DragMode.NoDrag); self.view.setCursor(Qt.CursorShape.CrossCursor); self.statusBar().showMessage("エッジ接続: 1つ目のノードをクリック")
            self.scene.update_preview_node(None, tool_name)
        elif tool_name == "paste":
            self.view.setDragMode(QGraphicsView.DragMode.NoDrag); self.view.setCursor(Qt.CursorShape.CrossCursor); self.statusBar().showMessage("ペーストモード")
            g_pos = QCursor.pos(); v_pos = self.view.mapFromGlobal(g_pos)
            if self.view.rect().contains(v_pos): self.scene.update_preview_node(self.view.mapToScene(v_pos), tool_name)
        else: 
            self.view.setDragMode(QGraphicsView.DragMode.NoDrag); self.view.setCursor(Qt.CursorShape.CrossCursor); self.statusBar().showMessage(f"配置モード: {tool_name}")
            g_pos = QCursor.pos(); v_pos = self.view.mapFromGlobal(g_pos)
            if self.view.rect().contains(v_pos): self.scene.update_preview_node(self.view.mapToScene(v_pos), tool_name)

    def on_selection_changed(self):
        sel = self.scene.selectedItems()
        nodes = [i for i in sel if isinstance(i, NodeItem)]
        if nodes:
            node = nodes[0]
            self.sb_node_w.blockSignals(True)
            self.sb_node_h.blockSignals(True)
            self.sb_node_w.setValue(node.w)
            self.sb_node_h.setValue(node.h)
            self.sb_node_w.blockSignals(False)
            self.sb_node_h.blockSignals(False)
            self.sb_node_w.setEnabled(True)
            self.sb_node_h.setEnabled(True)
        else:
            self.sb_node_w.setEnabled(False)
            self.sb_node_h.setEnabled(False)

    def on_node_size_ui_changed(self):
        nodes = [i for i in self.scene.selectedItems() if isinstance(i, NodeItem)]
        if nodes:
            w, h = self.sb_node_w.value(), self.sb_node_h.value()
            changed = False
            for n in nodes:
                if n.w != w or n.h != h:
                    n.set_size(w, h)
                    changed = True
            if changed:
                self.push_undo_state("ノードサイズ変更")

    def auto_layout_networkx(self):
        data = self.get_scene_json()
        if not data["nodes"]: return

        G = nx.DiGraph()
        for n in data["nodes"]: G.add_node(n["id"])
        for e in data["edges"]: G.add_edge(e["source"], e["target"])

        try:
            for layer, nodes in enumerate(nx.topological_generations(G) if nx.is_directed_acyclic_graph(G) else [list(G.nodes)]):
                for node in nodes:
                    G.nodes[node]["layer"] = layer
            pos = nx.multipartite_layout(G, subset_key="layer", align="horizontal")
            
            x_scale, y_scale = 200, 150
            for i in self.scene.items():
                if isinstance(i, NodeItem) and i.node_id in pos:
                    p = pos[i.node_id]
                    sx, sy = round(p[0]*x_scale/GRID_SIZE)*GRID_SIZE, round(p[1]*y_scale/GRID_SIZE)*GRID_SIZE
                    i.setPos(sx, sy)
            self.push_undo_state("自動レイアウト")
            QMessageBox.information(self, "完了", "自動レイアウトが完了しました。")
        except Exception as e:
            QMessageBox.warning(self, "エラー", f"自動レイアウトに失敗しました。\n{e}")

    def link_dxf_template(self):
        path, _ = QFileDialog.getOpenFileName(self, "DXF出力用テンプレート指定", "", "DXF Files (*.dxf)")
        if not path: return
        self.dxf_template_path = path
        self.update_dxf_template_label()
        self.load_dxf_frame(path)
        self.sb_frame_scale.setEnabled(True)
        self.push_undo_state("DXF出力用テンプレート指定")
        QMessageBox.information(self, "完了", f"テンプレートをリンクし、背景に表示しました。\n{path}")

    def clear_dxf_template(self):
        if self.dxf_template_path:
            self.dxf_template_path = None
            self.update_dxf_template_label()
            self.push_undo_state("DXF出力用テンプレート解除")
            QMessageBox.information(self, "完了", "設定したテンプレートリンクを解除しました。")
            if self.frame_item:
                self.scene.removeItem(self.frame_item)
                self.frame_item = None
            self.act_lock_frame.setEnabled(False)
            self.sb_frame_scale.setValue(1.0)
            self.sb_frame_scale.setEnabled(False)

    def load_dxf_frame(self, path, pos=QPointF(0, 0), scale=1.0):
        if self.frame_item:
            self.scene.removeItem(self.frame_item)
            self.frame_item = None
            
        items = import_dxf_as_items(path)
        if items:
            self.frame_item = DxfFrameItem(items)
            self.scene.addItem(self.frame_item)
            self.frame_item.setPos(pos)
            self.frame_item.setScale(scale)
            self.act_lock_frame.setEnabled(True)
            self.act_lock_frame.setChecked(True)
            self.sb_frame_scale.setEnabled(True)
            self.sb_frame_scale.blockSignals(True)
            self.sb_frame_scale.setValue(scale)
            self.sb_frame_scale.blockSignals(False)
            self.frame_item.set_locked(True)
            self.update_frame_lock_icon()
        else:
            QMessageBox.warning(self, "警告", "DXFの読み込みに失敗したか、有効なエンティティがありませんでした。")

    def change_frame_scale(self):
        if self.frame_item:
            self.dxf_template_scale = self.sb_frame_scale.value()
            self.frame_item.setScale(self.dxf_template_scale)
            self.push_undo_state("図面枠の倍率変更")

    def toggle_frame_lock(self):
        if self.frame_item:
            locked = self.act_lock_frame.isChecked()
            self.frame_item.set_locked(locked)
            self.update_frame_lock_icon()
            msg = "図面枠をロックしました。" if locked else "図面枠のロックを解除しました。ドラッグで移動できます。"
            self.statusBar().showMessage(msg, 3000)

    def update_frame_lock_icon(self):
        locked = self.act_lock_frame.isChecked()
        icon = 'fa5s.lock' if locked else 'fa5s.lock-open'
        text = "図面枠をロック中" if locked else "図面枠を移動可能"
        self.act_lock_frame.setText(text)
        # アイコンを更新（apply_themeと同様のロジック）
        icon_color = '#212529' if self.is_light_theme else '#F8F9FA'
        self.act_lock_frame.setIcon(qta.icon(icon, color=icon_color))
        # 既存のicon_actionsリストも更新しておく
        for i, (act, _) in enumerate(self.icon_actions):
            if act == self.act_lock_frame:
                self.icon_actions[i] = (act, icon)
                break

    def update_dxf_template_label(self):
        if hasattr(self, 'lbl_dxf_template'):
            if self.dxf_template_path:
                name = os.path.basename(self.dxf_template_path)
                self.lbl_dxf_template.setText(f" 図枠リンク中: {name} ")
                self.lbl_dxf_template.setStyleSheet("background-color: #E3F2FD; color: #1565C0; font-weight: bold; padding: 2px; border-radius: 4px;")
            else:
                self.lbl_dxf_template.setText(" 図枠リンク: [未設定] ")
                self.lbl_dxf_template.setStyleSheet("color: gray;")

    def generate_excel(self):
        default_name = datetime.datetime.now().strftime("%Y%m%d")
        path, _ = QFileDialog.getSaveFileName(self, "仕様書(Excel)生成", f"{default_name}.xlsx", "Excel Files (*.xlsx)")
        if not path: return
        
        data = self.get_scene_json()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Flowchart Specifications"
        
        headers = ["Step", "Node ID", "Type", "Text", "Next Steps"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True)

        id_to_node = {n["id"]: n for n in data["nodes"]}
        edges_from = {}
        for e in data["edges"]:
            edges_from.setdefault(e["source"], []).append(e["target"])

        row = 2
        for i, n in enumerate(data["nodes"], 1):
            next_ids = edges_from.get(n["id"], [])
            next_texts = [id_to_node[nid]["text"].replace('\n', ' ') for nid in next_ids if nid in id_to_node]
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=n["id"])
            ws.cell(row=row, column=3, value=n["type"])
            ws.cell(row=row, column=4, value=n["text"])
            ws.cell(row=row, column=5, value=", ".join(next_texts))
            row += 1

        try:
            wb.save(path)
            QMessageBox.information(self, "完了", f"仕様書を生成しました:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"保存に失敗しました:\n{e}")

    def import_drawio(self):
        path, _ = QFileDialog.getOpenFileName(self, "Draw.io インポート", "", "XML Files (*.xml *.drawio)")
        if not path: return
        try:
            tree = ET.parse(path)
            root = tree.getroot()
            
            self.scene.clear(); self.scene.items_ref.clear()
            id_map = {}
            for cell in root.iter('mxCell'):
                cid = cell.get('id')
                geom = cell.find('mxGeometry')
                if geom is None: continue
                
                if cell.get('edge') == '1':
                    src, tgt = cell.get('source'), cell.get('target')
                    if src in id_map and tgt in id_map:
                        style = cell.get('style', '')
                        arrow_style = "none" if "endArrow=none" in style else "end"
                        edge = EdgeItem(id_map[src], id_map[tgt], cell.get('value', ''), 2, "solid", "orthogonal", arrow_style)
                        self.scene.items_ref.append(edge); self.scene.addItem(edge); edge.update_position()
                elif cell.get('vertex') == '1':
                    x, y = float(geom.get('x', 0)), float(geom.get('y', 0))
                    txt = cell.get('value', '').replace('&lt;br&gt;', '\n').replace('<br>', '\n')
                    style = cell.get('style', '')
                    ntype = "process"
                    if "rhombus" in style: ntype = "decision"
                    elif "ellipse" in style: ntype = "terminal"
                    node = NodeItem(x, y, txt, ntype, cid)
                    self.scene.items_ref.append(node); self.scene.addItem(node); id_map[cid] = node

            self.last_state = self.get_scene_json()
            QMessageBox.information(self, "完了", "Draw.ioファイルのインポートが完了しました。")
        except Exception as e:
            QMessageBox.warning(self, "エラー", f"読み込みに失敗しました:\n{e}")

    def export_drawio(self):
        default_name = datetime.datetime.now().strftime("%Y%m%d")
        path, _ = QFileDialog.getSaveFileName(self, "Draw.io エクスポート", f"{default_name}.xml", "XML Files (*.xml)")
        if not path: return
        data = self.get_scene_json()
        
        mxfile = ET.Element('mxfile')
        diagram = ET.SubElement(mxfile, 'diagram', id=str(uuid.uuid4()), name="Page-1")
        mxGraphModel = ET.SubElement(diagram, 'mxGraphModel', dx="1000", dy="1000", grid="1", gridSize="20", guides="1", tooltips="1", connect="1", arrows="1", fold="1", page="1", pageScale="1", pageWidth="827", pageHeight="1169", math="0", shadow="0")
        root = ET.SubElement(mxGraphModel, 'root')
        ET.SubElement(root, 'mxCell', id="0")
        ET.SubElement(root, 'mxCell', id="1", parent="0")

        for n in data["nodes"]:
            style = "rounded=0;whiteSpace=wrap;html=1;"
            if n["type"] == "decision": style = "rhombus;whiteSpace=wrap;html=1;"
            elif n["type"] == "terminal": style = "ellipse;whiteSpace=wrap;html=1;"
            elif n["type"] == "data": style = "shape=parallelogram;perimeter=parallelogramPerimeter;whiteSpace=wrap;html=1;fixedSize=1;"
            
            cell = ET.SubElement(root, 'mxCell', id=n["id"], value=n["text"].replace('\n', '<br>'), style=style, vertex="1", parent="1")
            ET.SubElement(cell, 'mxGeometry', x=str(n["x"]-n["w"]/2), y=str(n["y"]-n["h"]/2), width=str(n["w"]), height=str(n["h"]), **{'as': 'geometry'})

        for i, e in enumerate(data["edges"]):
            style = "edgeStyle=orthogonalEdgeStyle;rounded=0;orthogonalLoop=1;jettySize=auto;html=1;" if e.get("routing")=="orthogonal" else "html=1;"
            arr = e.get("arrow", "end")
            if arr == "none": style += "startArrow=none;endArrow=none;"
            elif arr == "end": style += "startArrow=none;endArrow=classic;"
            elif arr == "start": style += "startArrow=classic;endArrow=none;"
            elif arr == "both": style += "startArrow=classic;endArrow=classic;"
            cell = ET.SubElement(root, 'mxCell', id=f"edge_{i}", value=e.get("label", ""), style=style, edge="1", parent="1", source=e["source"], target=e["target"])
            ET.SubElement(cell, 'mxGeometry', relative="1", **{'as': 'geometry'})

        tree = ET.ElementTree(mxfile)
        try:
            tree.write(path, encoding='utf-8', xml_declaration=True)
            QMessageBox.information(self, "完了", f"Draw.io形式でエクスポートしました。\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "エラー", str(e))

    def copy_items(self):
        sel_items = self.scene.selectedItems()
        if not sel_items: return
        
        self.clipboard_data = self.get_scene_json(selected_only=True)
        nodes_data = self.clipboard_data.get("nodes", [])
        if nodes_data:
            self.clipboard_base_pos = QPointF(min(n["x"] for n in nodes_data), min(n["y"] for n in nodes_data))
            self.statusBar().showMessage("コピーしました（クリックで配置）", 3000)
            
            self.scene.clearSelection()
            self.set_tool("paste")
        else:
            self.clipboard_base_pos = QPointF(0, 0)
            self.statusBar().showMessage("コピーに失敗しました", 3000)

    def paste_items(self):
        if self.clipboard_data and self.clipboard_data.get("nodes"): 
            self.set_tool("paste")

    def group_selected(self):
        items = [i for i in self.scene.selectedItems() if isinstance(i, NodeItem) and i.parentItem() is None]
        if len(items) < 2: return
        group = self.scene.createItemGroup(items); group.setFlags(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable | QGraphicsItem.GraphicsItemFlag.ItemIsMovable)
        self.scene.items_ref.append(group); self.push_undo_state("グループ化")

    def ungroup_selected(self):
        changed = False
        for item in self.scene.selectedItems():
            if type(item) == QGraphicsItemGroup:
                self.scene.destroyItemGroup(item)
                if item in self.scene.items_ref: self.scene.items_ref.remove(item)
                changed = True
        if changed: self.push_undo_state("グループ解除")

    def align_items(self, mode):
        nodes = [item for item in self.scene.selectedItems() if isinstance(item, NodeItem)]
        if len(nodes) < 2: return
        xs = [n.scenePos().x() for n in nodes]; ys = [n.scenePos().y() for n in nodes]
        if mode == "left": val = min(xs); [n.setPos(val, n.scenePos().y()) for n in nodes]
        elif mode == "right": val = max(xs); [n.setPos(val, n.scenePos().y()) for n in nodes]
        elif mode == "center_x": val = sum(xs)/len(xs); [n.setPos(val, n.scenePos().y()) for n in nodes]
        elif mode == "top": val = min(ys); [n.setPos(n.scenePos().x(), val) for n in nodes]
        elif mode == "bottom": val = max(ys); [n.setPos(n.scenePos().x(), val) for n in nodes]
        elif mode == "center_y": val = sum(ys)/len(ys); [n.setPos(n.scenePos().x(), val) for n in nodes]
        elif mode == "dist_h":
            nodes.sort(key=lambda n: n.scenePos().x()); span = (max(xs) - min(xs)) / (len(nodes) - 1)
            for i, n in enumerate(nodes): n.setPos(min(xs) + i*span, n.scenePos().y())
        elif mode == "dist_v":
            nodes.sort(key=lambda n: n.scenePos().y()); span = (max(ys) - min(ys)) / (len(nodes) - 1)
            for i, n in enumerate(nodes): n.setPos(n.scenePos().x(), min(ys) + i*span)
        self.push_undo_state(f"整列")

    def change_bg_color(self):
        nodes = [i for i in self.scene.selectedItems() if isinstance(i, NodeItem)]
        if not nodes: return
        c = QColorDialog.getColor(nodes[0].bg_color, self, "背景色")
        if c.isValid(): [n.set_bg_color(c) for n in nodes]; self.push_undo_state("背景色変更")

    def change_text_color(self):
        nodes = [i for i in self.scene.selectedItems() if isinstance(i, NodeItem)]
        if not nodes: return
        c = QColorDialog.getColor(nodes[0].text_color, self, "文字色")
        if c.isValid(): [n.set_text_color(c) for n in nodes]; self.push_undo_state("文字色変更")

    def change_line_color(self):
        items = self.scene.selectedItems()
        target_items = [i for i in items if isinstance(i, (NodeItem, EdgeItem))]
        if not target_items: return
        initial_color = target_items[0].line_color if target_items[0].line_color else QColor(Qt.GlobalColor.black)
        c = QColorDialog.getColor(initial_color, self, "線の色")
        if c.isValid():
            for i in target_items: i.set_line_color(c)
            self.push_undo_state("線の色変更")

    def change_edge_style(self):
        edges = [i for i in self.scene.selectedItems() if isinstance(i, EdgeItem)]
        if not edges: return
        w = int(self.cb_width.currentText()); s = self.cb_style.currentText().split("(")[1].replace(")","")
        r = self.cb_routing.currentData()
        a = self.cb_arrow.currentData()
        for e in edges: e.line_width = w; e.line_style = s; e.routing = r; e.arrow = a; e.update_pen(); e.update_position()
        self.push_undo_state("線のスタイル変更")

    def delete_selected_items(self):
        sel = self.scene.selectedItems()
        if not sel: return
        edges, nodes, wps = set(), set(), set()
        for i in sel:
            if type(i) == QGraphicsItemGroup: 
                self.ungroup_selected(); return 
            if isinstance(i, NodeItem): nodes.add(i); edges.update(i.edges)
            elif isinstance(i, EdgeItem): edges.add(i)
            elif isinstance(i, WaypointItem): wps.add(i)
            
        for wp in wps:
            if wp.edge not in edges: wp.edge.remove_waypoint(wp)
            
        for e in edges:
            for wp in e.waypoints:
                if wp.scene(): self.scene.removeItem(wp)
                if wp in self.scene.items_ref: self.scene.items_ref.remove(wp)
            e.waypoints.clear()
            if e in e.source_node.edges: e.source_node.edges.remove(e)
            if e in e.target_node.edges: e.target_node.edges.remove(e)
            if e.scene(): self.scene.removeItem(e)
            if e in self.scene.items_ref: self.scene.items_ref.remove(e)
            
        for n in nodes:
            if n.scene(): self.scene.removeItem(n)
            if n in self.scene.items_ref: self.scene.items_ref.remove(n)
            
        for bg in bgs:
            if bg.scene(): self.scene.removeItem(bg)
            if bg in self.scene.items_ref: self.scene.items_ref.remove(bg)
            
        self.push_undo_state("削除")

    def change_font_family(self):
        items = self.scene.selectedItems()
        family = self.cb_font.currentText()
        for item in items:
            if hasattr(item, 'set_font_family'):
                item.set_font_family(family)
        self.push_undo_state("フォント変更")

    def save_file(self):
        if self.current_filepath:
            with open(self.current_filepath, 'w', encoding='utf-8') as f: json.dump(self.get_scene_json(), f, indent=4, ensure_ascii=False)
            self.statusBar().showMessage("上書き保存しました", 5000)
        else: self.save_file_as()

    def save_file_as(self):
        default_name = datetime.datetime.now().strftime("%Y%m%d")
        path, _ = QFileDialog.getSaveFileName(self, "名前を付けて保存", f"{default_name}.json", "JSON Files (*.json)")
        if path: self.current_filepath = path; self.save_file(); self.update_window_title()

    def load_json(self):
        path, _ = QFileDialog.getOpenFileName(self, "読込", "", "JSON Files (*.json)")
        if path:
            with open(path, 'r', encoding='utf-8') as f: data = json.load(f)
            self.load_scene_json(data); self.undo_stack.clear(); self.last_state = self.get_scene_json(); self.current_filepath = path; self.update_window_title()

    def export_file(self, initial_filter=None):
        default_name = datetime.datetime.now().strftime("%Y%m%d")
        all_filters = "DXF Files (*.dxf);;PDF Files (*.pdf);;JPEG Files (*.jpeg *.jpg);;PNG Files (*.png);;SVG Files (*.svg);;Mermaid Markdown (*.md)"
        path, chosen_filter = QFileDialog.getSaveFileName(self, "エクスポート", default_name, all_filters, initialFilter=initial_filter if initial_filter else "")
        if not path: return
        self.scene.clearSelection(); rect = self.scene.itemsBoundingRect().adjusted(-20, -20, 20, 20)
        hidden = []
        for i in self.scene.items():
            try:
                if (isinstance(i, WaypointItem) or i == getattr(self.scene, 'preview_node', None) or i in getattr(self.scene, 'preview_items', [])) and i.isVisible():
                    i.hide(); hidden.append(i)
            except RuntimeError: pass
        old_grid = self.scene.draw_grid
        self.scene.draw_grid = False
        if not rect.isEmpty():
            if path.endswith(('.png', '.jpg', '.jpeg')):
                img = QImage(rect.size().toSize(), QImage.Format.Format_ARGB32); img.fill(Qt.GlobalColor.white)
                p = QPainter(img); self.scene.render(p, QRectF(img.rect()), rect); p.end(); img.save(path)
            elif path.endswith('.svg'):
                gen = QSvgGenerator(); gen.setFileName(path); gen.setSize(rect.size().toSize()); gen.setViewBox(rect)
                p = QPainter(gen); self.scene.render(p, QRectF(0, 0, rect.width(), rect.height()), rect); p.end()
            elif path.endswith('.pdf'):
                printer = QPrinter(QPrinter.PrinterMode.HighResolution)
                printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
                printer.setOutputFileName(path)
                layout = QPageLayout(QPageSize(rect.size(), QPageSize.Unit.Point), QPageLayout.Orientation.Portrait, QMarginsF(0, 0, 0, 0))
                printer.setPageLayout(layout)
                p = QPainter(printer)
                self.scene.render(p, QRectF(printer.pageRect(QPrinter.Unit.DevicePixel)), rect)
                p.end()
            elif path.endswith('.dxf'): 
                export_dxf_file(self.scene, path, self.cb_dxf_ver.currentData(), self.dxf_template_path)
            elif path.endswith('.md'): self._export_mermaid(path)
        self.scene.draw_grid = old_grid
        for i in hidden: i.show()
        QMessageBox.information(self, "完了", f"エクスポート完了:\n{path}")

    def _export_mermaid(self, path):
        data = self.get_scene_json()
        lines = ["```mermaid", "graph TD"]
        for n in data.get("nodes", []):
            nid = n["id"].replace("-", "")
            text = n["text"].replace("\n", "<br>")
            shape_open, shape_close = ("[", "]")
            if n["type"] == "decision": shape_open, shape_close = ("{", "}")
            elif n["type"] == "terminal": shape_open, shape_close = ("([", "])")
            elif n["type"] == "data": shape_open, shape_close = ("[/", "/]")
            lines.append(f'    {nid}{shape_open}"{text}"{shape_close}')
        for e in data.get("edges", []):
            src = e["source"].replace("-", "")
            tgt = e["target"].replace("-", "")
            label = e.get("label", "").replace("\n", "<br>")
            arr_type = e.get("arrow", "end")
            style_type = e.get("style", "solid")
            is_dotted = style_type in ["dash", "dot"]
            
            if arr_type == "both": arrow = "<-.->" if is_dotted else "<->"
            elif arr_type == "start": arrow = "<-.-" if is_dotted else "<--"
            elif arr_type == "end": arrow = "-.->" if is_dotted else "-->"
            else: arrow = "-.-" if is_dotted else "---"
            
            if label: lines.append(f'    {src} {arrow}|"{label}"| {tgt}')
            else: lines.append(f'    {src} {arrow} {tgt}')
        lines.append("```")
        with open(path, "w", encoding="utf-8") as f: f.write("\n".join(lines))

    def copy_to_jwcad(self):
        td = ["JwcTemp", "hq", "lc7", "lt1"]
        for item in self.scene.items():
            is_preview = False
            try:
                if item == getattr(self.scene, 'preview_node', None) or item in getattr(self.scene, 'preview_items', []): is_preview = True
            except RuntimeError: pass
            if is_preview: continue
            
            if isinstance(item, NodeItem):
                x, y = item.scenePos().x(), -item.scenePos().y(); t = item.node_type
                hw, hh = item.w / 2, item.h / 2
                def add_p(ps):
                    for i in range(len(ps)): td.append(f"{ps[i][0]} {ps[i][1]} {ps[(i+1)%len(ps)][0]} {ps[(i+1)%len(ps)][1]}")
                if t == "process": add_p([(x-hw, y-hh), (x+hw, y-hh), (x+hw, y+hh), (x-hw, y+hh)])
                elif t == "decision": 
                    dw, dh = hw + 10, hh + 10
                    add_p([(x, y-dh), (x+dw, y), (x, y+dh), (x-dw, y)])
                elif t == "data":
                    skew = hh
                    add_p([(x-hw+skew/2, y-hh), (x+hw+skew/2, y-hh), (x+hw-skew/2, y+hh), (x-hw-skew/2, y+hh)])
                elif t == "terminal":
                    add_p([(x-hw+hh, y-hh), (x+hw-hh, y-hh), (x+hw, y-hh/2), (x+hw, y+hh/2), (x+hw-hh, y+hh), (x-hw+hh, y+hh), (x-hw, y+hh/2), (x-hw, y-hh/2)])
                else: add_p([(x-hw, y-hh), (x+hw, y-hh), (x+hw, y+hh), (x-hw, y+hh)])
                if item.text_item.toPlainText():
                    ls = item.text_item.toPlainText().split('\n'); sy = y + (len(ls)-1)*7.5
                    for i, l in enumerate(ls): wc = sum(2 if unicodedata.east_asian_width(c) in 'FWA' else 1 for c in l); td.append(f'ch {x-wc*2.5} {sy-i*15-6.0} 10 0 "{l}')
            elif isinstance(item, EdgeItem):
                pts = [item.source_node.scenePos()] + [wp.scenePos() for wp in item.waypoints] + [item.target_node.scenePos()]
                if len(pts) >= 2: pts[0] = clip_line_to_node(pts[0], pts[1], item.source_node); pts[-1] = clip_line_to_node(pts[-1], pts[-2], item.target_node)
                for i in range(len(pts)-1): td.append(f"{pts[i].x()} {-pts[i].y()} {pts[i+1].x()} {-pts[i+1].y()}")
                if item.raw_text:
                    pos = item.get_auto_text_pos() + (item.text_item.manual_offset if item.text_item.manual_offset else QPointF(0,0))
                    mx, my = pos.x() + item.text_item.boundingRect().width()/2, -(pos.y() + item.text_item.boundingRect().height()/2)
                    ls = item.raw_text.split('\n'); sy = my + (len(ls)-1)*6
                    for i, l in enumerate(ls): wc = sum(2 if unicodedata.east_asian_width(c) in 'FWA' else 1 for c in l); td.append(f'ch {mx-wc*2.5} {sy-i*12-5.0} 10 0 "{l}')
                if item.arrow == "end" and hasattr(item, 'arrow_p1') and hasattr(item, 'arrow_p2'):
                    angle = math.atan2(item.arrow_p2.y() - item.arrow_p1.y(), item.arrow_p2.x() - item.arrow_p1.x())
                    arrow_size = 12 + item.line_width * 1.5
                    
                    arrow_p1 = item.arrow_p2 - QPointF(math.cos(angle + math.pi / 6) * arrow_size, math.sin(angle + math.pi / 6) * arrow_size)
                    arrow_p2 = item.arrow_p2 - QPointF(math.cos(angle - math.pi / 6) * arrow_size, math.sin(angle - math.pi / 6) * arrow_size)
                    td.append(f"{item.arrow_p2.x()} {-item.arrow_p2.y()} {arrow_p1.x()} {-arrow_p1.y()}")
                    td.append(f"{arrow_p1.x()} {-arrow_p1.y()} {arrow_p2.x()} {-arrow_p2.y()}")
                    td.append(f"{arrow_p2.x()} {-arrow_p2.y()} {item.arrow_p2.x()} {-item.arrow_p2.y()}")
        QApplication.clipboard().setText('\r\n'.join(td) + '\r\n')
        QMessageBox.information(self, "完了", "Jw_cad用のデータをクリップボードにコピーしました。\nJw_cadを開いて「編集」→「貼り付け (Ctrl+V)」を実行してください。")

    def open_print_dialog(self): 
        CustomPrintPreviewDialog(self, len(self.scene.selectedItems()) > 0).exec()

if __name__ == '__main__':
    if sys.platform == 'win32':
        import ctypes
        myappid = 'tsukasamiyashita.dxfflowchartmiya.v1.0.0'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

    app = QApplication(sys.argv)
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec())